#!/usr/bin/env python3
"""
CSV to XLSX Google Drive Manager for Hebrew Translation (API v3)
Converts CSV files to XLSX format and manages them on Google Drive using official Google Drive API v3
"""

import argparse
import os
import sys
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Google Drive API v3 imports
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload

# If modifying these scopes, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/drive.file']

class CSVXLSXDriveManager:
    def __init__(self):
        self.service = None
        
    def authenticate(self):
        """Authenticate with Google Drive API using official method"""
        creds = None
        # The file token.json stores the user's access and refresh tokens, and is
        # created automatically when the authorization flow completes for the first time.
        if os.path.exists('token.json'):
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        
        # If there are no (valid) credentials available, let the user log in.
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                if not os.path.exists('credentials.json'):
                    print("❌ credentials.json file not found!")
                    print("\nTo set up Google Drive API:")
                    print("1. Go to: https://console.cloud.google.com/")
                    print("2. Create a new project or select existing one")
                    print("3. Enable Google Drive API")
                    print("4. Create credentials (OAuth 2.0 Client IDs) - Desktop Application")
                    print("5. Download credentials.json to this directory")
                    return False
                    
                flow = InstalledAppFlow.from_client_secrets_file(
                    'credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            
            # Save the credentials for the next run
            with open('token.json', 'w') as token:
                token.write(creds.to_json())
        
        try:
            self.service = build('drive', 'v3', credentials=creds)
            print("✅ Successfully authenticated with Google Drive")
            return True
        except Exception as e:
            print(f"❌ Authentication failed: {e}")
            return False

    def csv_to_xlsx(self, csv_file_path, xlsx_file_path, sheet_name="Translation", use_checkboxes=False):
        """Convert CSV to XLSX with proper formatting
        
        Args:
            csv_file_path: Path to input CSV file
            xlsx_file_path: Path to output XLSX file  
            sheet_name: Name of the Excel sheet
            use_checkboxes: If True, use actual checkboxes instead of dropdown (experimental)
        """
        try:
            # Try to read CSV with different encodings (UTF-8 first, then Windows-1255)
            df = None
            for encoding in ['utf-8', 'windows-1255', 'utf-8-sig']:
                try:
                    # Read CSV without headers and assign proper column names
                    df = pd.read_csv(csv_file_path, encoding=encoding, header=None, 
                                   names=['english', 'hebrew', 'tested?', 'comments'])
                    print(f"✅ Successfully read CSV with {encoding} encoding")
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                raise ValueError("Could not read CSV file with any supported encoding")
            
            # Fill any missing data with empty strings
            df = df.fillna('')
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Import data validation
                from openpyxl.worksheet.datavalidation import DataValidation
                
                # Set column widths for better readability
                column_widths = {
                    'A': 50,   # english
                    'B': 50,   # hebrew  
                    'C': 15,   # tested (boolean dropdown)
                    'D': 100,  # comments
                }
                
                for column, width in column_widths.items():
                    worksheet.column_dimensions[column].width = width
                
                # Add data validation for the "tested" column (column C)
                # Create dropdown with TRUE/FALSE options
                dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
                dv.error = 'Your entry is not valid'
                dv.errorTitle = 'Invalid Entry'
                dv.prompt = 'Please select TRUE or FALSE'
                dv.promptTitle = 'Testing Status'
                
                # Apply validation to the entire "tested" column (starting from row 2, skipping header)
                worksheet.add_data_validation(dv)
                dv.add(f"C2:C{len(df) + 1}")
                
                # Add conditional formatting for TRUE/FALSE values
                from openpyxl.formatting.rule import FormulaRule
                from openpyxl.styles import PatternFill, Font
                
                # Green background for TRUE
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                true_rule = FormulaRule(formula=['$C2="TRUE"'], fill=green_fill)
                worksheet.conditional_formatting.add(f"C2:C{len(df) + 1}", true_rule)
                
                # Light red background for FALSE  
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                false_rule = FormulaRule(formula=['$C2="FALSE"'], fill=red_fill)
                worksheet.conditional_formatting.add(f"C2:C{len(df) + 1}", false_rule)
                
                # Set text alignment for different columns
                for row in range(2, len(df) + 2):  # Skip header row
                    # English text column (A)
                    cell_a = worksheet[f'A{row}']
                    cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    
                    # Hebrew translation column (B)
                    cell_b = worksheet[f'B{row}']
                    cell_b.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
                    
                    # Tested column (C) - center alignment for better readability
                    cell_c = worksheet[f'C{row}']
                    cell_c.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Comments column (D) 
                    cell_d = worksheet[f'D{row}']
                    cell_d.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            print(f"✅ Successfully converted {csv_file_path} to {xlsx_file_path}")
            return True
            
        except Exception as e:
            print(f"❌ Error converting CSV to XLSX: {e}")
            return False

    def upload_to_drive(self, file_path, title=None, folder_id=None):
        """Upload file to Google Drive"""
        if not self.service:
            if not self.authenticate():
                return None
        
        try:
            file_metadata = {
                'name': title or os.path.basename(file_path)
            }
            
            if folder_id:
                file_metadata['parents'] = [folder_id]
            
            media = MediaFileUpload(file_path, resumable=True)
            
            file = self.service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id,name,webViewLink'
            ).execute()
            
            print(f"✅ Successfully uploaded '{file.get('name')}' to Google Drive")
            print(f"📂 File ID: {file.get('id')}")
            print(f"🔗 View Link: {file.get('webViewLink')}")
            
            return file.get('id')
            
        except HttpError as error:
            print(f"❌ Error uploading to Google Drive: {error}")
            return None

    def download_from_drive(self, file_id, output_path):
        """Download file from Google Drive"""
        if not self.service:
            if not self.authenticate():
                return False
        
        try:
            # Get file metadata
            file_metadata = self.service.files().get(fileId=file_id).execute()
            print(f"📥 Downloading: {file_metadata.get('name')}")
            
            # Download file content
            request = self.service.files().get_media(fileId=file_id)
            with open(output_path, 'wb') as f:
                downloader = MediaIoBaseDownload(f, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()
                    print(f"Download progress: {int(status.progress() * 100)}%")
            
            print(f"✅ Successfully downloaded to {output_path}")
            return True
            
        except HttpError as error:
            print(f"❌ Error downloading from Google Drive: {error}")
            return False

    def list_files(self, folder_id=None):
        """List files in Google Drive"""
        if not self.service:
            if not self.authenticate():
                return []
        
        try:
            query = "trashed=false"
            if folder_id:
                query += f" and parents in '{folder_id}'"
            
            results = self.service.files().list(
                q=query,
                pageSize=50,
                fields="nextPageToken, files(id, name, createdTime, modifiedTime, webViewLink)"
            ).execute()
            
            items = results.get('files', [])
            
            if not items:
                print("📁 No files found.")
                return []
            
            print("📁 Files in Google Drive:")
            for item in items:
                print(f"  📄 {item['name']} (ID: {item['id']})")
                print(f"     🔗 {item.get('webViewLink', 'N/A')}")
                print()
            
            return items
            
        except HttpError as error:
            print(f"❌ Error listing files: {error}")
            return []

    def delete_file(self, file_id):
        """Delete a file from Google Drive"""
        if not self.service:
            if not self.authenticate():
                return False
        
        try:
            # Get file metadata first to show what we're deleting
            file_metadata = self.service.files().get(fileId=file_id).execute()
            file_name = file_metadata.get('name')
            
            # Confirm deletion
            response = input(f"⚠️  Are you sure you want to delete '{file_name}'? (y/N): ")
            if response.lower() != 'y':
                print("❌ Deletion cancelled")
                return False
            
            # Delete the file
            self.service.files().delete(fileId=file_id).execute()
            print(f"✅ Successfully deleted '{file_name}' from Google Drive")
            return True
            
        except HttpError as error:
            print(f"❌ Error deleting file: {error}")
            return False

    def xlsx_to_csv(self, xlsx_file_path, csv_file_path, sheet_name="Translation"):
        """Convert XLSX to CSV"""
        try:
            # Read Excel file
            df = pd.read_excel(xlsx_file_path, sheet_name=sheet_name)
            
            # Save as CSV with UTF-8 encoding
            df.to_csv(csv_file_path, index=False, encoding='utf-8')
            
            print(f"✅ Successfully converted {xlsx_file_path} to {csv_file_path}")
            return True
            
        except Exception as e:
            print(f"❌ Error converting XLSX to CSV: {e}")
            return False

def main():
    parser = argparse.ArgumentParser(
        description="CSV to XLSX Google Drive Manager for Hebrew Translation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Upload CSV as XLSX to Google Drive
  %(prog)s --upload messages.csv --title "KQ1 Hebrew Messages"
  
  # Download XLSX from Google Drive as CSV
  %(prog)s --download --file-id 1ABC...XYZ --output messages.csv
  
  # Delete file from Google Drive
  %(prog)s --delete 1ABC...XYZ
  
  # List files in Google Drive
  %(prog)s --list
  
  # Convert CSV to XLSX locally (no upload)
  %(prog)s --csv-to-xlsx messages.csv messages.xlsx
  
  # Convert XLSX to CSV locally (no download)
  %(prog)s --xlsx-to-csv messages.xlsx messages.csv
        """
    )
    
    # Main operation (mutually exclusive)
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--upload', help='CSV file to upload to Google Drive')
    group.add_argument('--download', action='store_true', help='Download XLSX from Google Drive and convert to CSV')
    group.add_argument('--delete', help='Delete file from Google Drive (provide file ID)')
    group.add_argument('--list', action='store_true', help='List files in Google Drive')
    group.add_argument('--csv-to-xlsx', nargs=2, metavar=('CSV_FILE', 'XLSX_FILE'), help='Convert CSV to XLSX locally')
    group.add_argument('--xlsx-to-csv', nargs=2, metavar=('XLSX_FILE', 'CSV_FILE'), help='Convert XLSX to CSV locally')
    
    # Optional arguments
    parser.add_argument('--title', help='Title for uploaded file (default: filename)')
    parser.add_argument('--file-id', help='Google Drive file ID for download')
    parser.add_argument('--output', help='Output CSV file path for download')
    parser.add_argument('--folder-id', help='Google Drive folder ID for upload')
    parser.add_argument('--sheet-name', default='Translation', help='Excel sheet name (default: Translation)')
    
    args = parser.parse_args()
    
    manager = CSVXLSXDriveManager()
    
    try:
        if args.upload:
            # Upload CSV as XLSX to Google Drive
            if not os.path.exists(args.upload):
                print(f"❌ File not found: {args.upload}")
                sys.exit(1)
            
            # Convert CSV to XLSX in temp file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_xlsx = temp_file.name
            
            if manager.csv_to_xlsx(args.upload, temp_xlsx, args.sheet_name):
                file_id = manager.upload_to_drive(temp_xlsx, args.title, args.folder_id)
                if file_id:
                    print(f"🎉 Upload completed successfully!")
                else:
                    sys.exit(1)
            else:
                sys.exit(1)
            
            # Clean up temp file
            try:
                os.unlink(temp_xlsx)
            except:
                pass
                
        elif args.download:
            # Download XLSX from Google Drive and convert to CSV
            if not args.file_id:
                print("❌ --file-id is required for download")
                sys.exit(1)
            if not args.output:
                print("❌ --output is required for download")
                sys.exit(1)
            
            # Download to temp XLSX file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_xlsx = temp_file.name
            
            if manager.download_from_drive(args.file_id, temp_xlsx):
                if manager.xlsx_to_csv(temp_xlsx, args.output, args.sheet_name):
                    print(f"🎉 Download and conversion completed successfully!")
                else:
                    sys.exit(1)
            else:
                sys.exit(1)
            
            # Clean up temp file
            try:
                os.unlink(temp_xlsx)
            except:
                pass
                
        elif args.delete:
            # Delete file from Google Drive
            if manager.delete_file(args.delete):
                print("🎉 File deleted successfully!")
            else:
                sys.exit(1)
                
        elif args.list:
            # List files in Google Drive
            manager.list_files(args.folder_id)
            
        elif args.csv_to_xlsx:
            # Convert CSV to XLSX locally
            csv_file, xlsx_file = args.csv_to_xlsx
            if not os.path.exists(csv_file):
                print(f"❌ File not found: {csv_file}")
                sys.exit(1)
            
            if manager.csv_to_xlsx(csv_file, xlsx_file, args.sheet_name):
                print("🎉 Conversion completed successfully!")
            else:
                sys.exit(1)
                
        elif args.xlsx_to_csv:
            # Convert XLSX to CSV locally
            xlsx_file, csv_file = args.xlsx_to_csv
            if not os.path.exists(xlsx_file):
                print(f"❌ File not found: {xlsx_file}")
                sys.exit(1)
            
            if manager.xlsx_to_csv(xlsx_file, csv_file, args.sheet_name):
                print("🎉 Conversion completed successfully!")
            else:
                sys.exit(1)
    
    except KeyboardInterrupt:
        print("\n⏹️  Operation cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
